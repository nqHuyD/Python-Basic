# path = Path("ecommerce")
# path = Path("emails")
#
# # path.mkdir()
# path.rmdir()
# print(path.exists())

#path = path()

#for file in path.glob('*.py'):
#    print(file)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
	wb = xl.load_workbook(filename)
	sheet = wb['sheet1']
	for row in range(2, sheet.max_row + 1):
		cell =sheet.cell(row, 3)
		corrected_price = cell.value * .9
		correct_price_cell = sheet.cell(row, 4)
		correct_price_cell.value = corrected_price
	wb.save(f"{filename}22.xlsx")
	values = Reference(sheet,
		  min_row=2,
		  max_row=sheet.max_row,
		  min_col=4,
		  max_col=4)
process_workbook('transactions.xlsx')
